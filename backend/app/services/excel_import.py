from __future__ import annotations

import re
import zipfile
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook


class UnsafeWorkbook(ValueError):
    pass


CONTROL_CHARACTERS = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f\x7f]")


def clean_cell(value: object, *, limit: int = 2000) -> str:
    if value is None:
        return ""
    if isinstance(value, (date, datetime)):
        text = value.isoformat()
    else:
        text = str(value)
    return CONTROL_CHARACTERS.sub("", text).strip()[:limit]


def inspect_xlsx_archive(
    path: Path, *, max_uncompressed_bytes: int, max_members: int = 250
) -> None:
    try:
        with zipfile.ZipFile(path) as archive:
            members = archive.infolist()
            if len(members) > max_members:
                raise UnsafeWorkbook("Le classeur contient trop de fichiers internes.")
            total = 0
            for member in members:
                if member.flag_bits & 1:
                    raise UnsafeWorkbook("Les classeurs chiffrés ne sont pas acceptés.")
                total += member.file_size
                if total > max_uncompressed_bytes:
                    raise UnsafeWorkbook("Le classeur décompressé dépasse la taille autorisée.")
                if member.compress_size and member.file_size / member.compress_size > 200:
                    raise UnsafeWorkbook("Le taux de compression du classeur est excessif.")
            names = {member.filename for member in members}
            if "[Content_Types].xml" not in names or "xl/workbook.xml" not in names:
                raise UnsafeWorkbook("Le fichier ne contient pas un classeur XLSX valide.")
    except zipfile.BadZipFile as exc:
        raise UnsafeWorkbook("Le fichier XLSX est invalide.") from exc


def read_xlsx(
    path: Path,
    *,
    max_rows: int,
    max_columns: int,
    max_uncompressed_bytes: int,
) -> tuple[list[dict[str, object]], list[list[str]]]:
    inspect_xlsx_archive(path, max_uncompressed_bytes=max_uncompressed_bytes)
    workbook = load_workbook(path, read_only=True, data_only=True, keep_links=False)
    try:
        if not workbook.sheetnames:
            raise UnsafeWorkbook("Le classeur ne contient aucune feuille.")
        sheet = workbook[workbook.sheetnames[0]]
        if sheet.max_column > max_columns:
            raise UnsafeWorkbook(f"Le classeur dépasse la limite de {max_columns} colonnes.")
        iterator = sheet.iter_rows(values_only=True)
        header_values = next(iterator, None)
        if header_values is None:
            raise UnsafeWorkbook("Le classeur est vide.")
        headers = [clean_cell(value, limit=200) for value in header_values]
        while headers and not headers[-1]:
            headers.pop()
        if not headers or not any(headers):
            raise UnsafeWorkbook("La première ligne doit contenir des en-têtes.")
        columns = [
            {"index": index, "name": name or f"Colonne {index + 1}"}
            for index, name in enumerate(headers)
        ]
        rows: list[list[str]] = []
        for row_number, values in enumerate(iterator, start=2):
            if row_number > max_rows + 1:
                raise UnsafeWorkbook(f"Le classeur dépasse la limite de {max_rows} lignes.")
            row = [clean_cell(value) for value in values[: len(columns)]]
            row.extend([""] * (len(columns) - len(row)))
            if any(row):
                rows.append(row)
        if not rows:
            raise UnsafeWorkbook("Le classeur ne contient aucune ligne de données.")
        return columns, rows
    finally:
        workbook.close()
